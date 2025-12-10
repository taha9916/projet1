import openpyxl
from openpyxl.styles import PatternFill, Font
import logging
import pandas as pd
import re

logger = logging.getLogger(__name__)

# Mapping des paramètres environnementaux vers les lignes du fichier Excel SLRI
PARAMETER_ROW_MAPPING = {
    # Eau
    'Température': 'Température',
    'temperature': 'Température',
    'Ph': 'Ph',
    'ph': 'Ph',
    'pH': 'Ph',
    'Turbidité': 'Turbidité',
    'turbidite': 'Turbidité',
    'Conductivité': 'Conductivité',
    'conductivite': 'Conductivité',
    'DBO5': 'DBO5',
    'DCO': 'DCO',
    'Oxygène dissous': 'Oxygène dissous',
    'oxygene dissous': 'Oxygène dissous',
    'Nitrates (NO₃⁻)': 'Nitrates (NO₃⁻)',
    'Nitrates (NO3-)': 'Nitrates (NO₃⁻)',
    'Nitrites (NO₂⁻)': 'Nitrites (NO₂⁻)',
    'Nitrites (NO2-)': 'Nitrites (NO₂⁻)',
    'Ammoniac (NH₄⁺)': 'Ammoniac (NH₄⁺)',
    'Ammoniac (NH4+)': 'Ammoniac (NH₄⁺)',
    'Phosphore total': 'Phosphore total',
    'Azote total': 'Azote total',
    'Plomb (Pb)': 'Plomb (Pb)',
    'Cadmium (Cd)': 'Cadmium (Cd)',
    'Chrome total (Cr)': 'Chrome total (Cr)',
    'Cuivre (Cu)': 'Cuivre (Cu)',
    'Zinc (Zn)': 'Zinc (Zn)',
    'Nickel (Ni)': 'Nickel (Ni)',
    'Mercure (Hg)': 'Mercure (Hg)',
    'Arsenic (As)': 'Arsenic (As)',
    'Hydrocarbures (HCT, HAP)': 'Hydrocarbures (HCT, HAP)',
    # Sol
    'Perméabilité': 'Perméabilité',
    'permeabilite': 'Perméabilité',
    'Matière organique': 'Matière organique',
    'matiere organique': 'Matière organique',
    'Carbone organique': 'Carbone organique',
    'carbone organique': 'Carbone organique',
    # Air
    'Poussières totales': 'Poussières totales',
    'poussieres totales': 'Poussières totales',
    'PM10': 'PM10',
    'PM2.5': 'PM2.5',
    'SO₂': 'SO₂',
    'SO2': 'SO₂',
    'NOx': 'NOx',
    'CO': 'CO',
    'O₃ (ozone)': 'O₃ (ozone)',
    'O3 (ozone)': 'O₃ (ozone)',
    # Population/habitats
    'Résidentiel': 'Résidentiel',
    'residential': 'Résidentiel',
    'Radiations électromagnétiques éoliennes': 'Radiations électromagnétiques éoliennes',
    'Radiations electromagnetiques eoliennes': 'Radiations électromagnétiques éoliennes',
    'Radiations électromagnétiques CABLES': 'Radiations électromagnétiques CABLES',
    'Radiations electromagnetiques CABLES': 'Radiations électromagnétiques CABLES',
    'Radiations électromagnétiques ONDEURS': 'Radiations électromagnétiques ONDEURS',
    'Radiations electromagnetiques ONDEURS': 'Radiations électromagnétiques ONDEURS',
    'Poussières': 'Poussières',
    'poussieres': 'Poussières',
    'Risques électriques et incendie': 'Risques électriques et incendie',
    'risques electriques et incendie': 'Risques électriques et incendie',
}

# Phases SLRI et leurs feuilles correspondantes
SLRI_PHASES = {
    'pre_construction': ['PRE CONSTRUCTION', 'Pré-construction', 'Pre-construction', 'État initial'],
    'construction': ['CONSTRUCTION', 'Construction'],
    'exploitation': ['EXPLOITATION', 'Exploitation'],
    'demantelement': ['DÉMANTÈLEMENT', 'Démantèlement', 'Demantelement', 'Démantèlement']
}

import unicodedata

def remove_accents(input_str):
    """Supprime les accents d'une chaîne (unicode -> ascii)."""
    return ''.join(c for c in unicodedata.normalize('NFD', input_str) if unicodedata.category(c) != 'Mn')

def normalize_parameter_name(param_name):
    """Normalise le nom d'un paramètre pour le mapping (sans accent, insensible à la casse).
    Args:
        param_name (str): Nom du paramètre brut
    Returns:
        str: Nom normalisé ou None si non trouvé
    """
    if not param_name:
        return None

    # Nettoyer, enlever accents, minuscule
    normalized = remove_accents(param_name.lower().strip())
    normalized = re.sub(r'\([^)]*\)', '', normalized).strip()

    # Mapping enrichi avec variantes anglaises
    mapping = {}
    for k, v in PARAMETER_ROW_MAPPING.items():
        k_norm = remove_accents(k.lower())
        mapping[k_norm] = v
    # Ajout variantes anglaises (exemples)
    mapping.update({
        'temperature': 'Température',
        'turbidity': 'Turbidité',
        'conductivity': 'Conductivité',
        'dissolved oxygen': 'Oxygène dissous',
        'nitrate': 'Nitrates (NO₃⁻)',
        'nitrite': 'Nitrites (NO₂⁻)',
        'ammonia': 'Ammoniac (NH₄⁺)',
        'lead': 'Plomb (Pb)',
        'cadmium': 'Cadmium (Cd)',
        'chromium': 'Chrome total (Cr)',
        'copper': 'Cuivre (Cu)',
        'zinc': 'Zinc (Zn)',
        'nickel': 'Nickel (Ni)',
        'mercury': 'Mercure (Hg)',
        'arsenic': 'Arsenic (As)',
        'organic matter': 'Matière organique',
        'permeability': 'Perméabilité',
        'dust': 'Poussières',
        'pm10': 'PM10',
        'pm2.5': 'PM2.5',
        'so2': 'SO₂',
        'nox': 'NOx',
        'co': 'CO',
        'ozone': 'O₃ (ozone)',
        'ph': 'Ph',
        'dbo5': 'DBO5',
        'dco': 'DCO',
    })

    # Recherche stricte
    if normalized in mapping:
        return mapping[normalized]
    # Recherche partielle
    for key, value in mapping.items():
        if key in normalized or normalized in key:
            return value
    return None


def find_parameter_row(ws, param_name, start_row=1, end_row=100):
    """Trouve la ligne d'un paramètre spécifique dans la feuille.
    Ajoute un log détaillé pour chaque tentative de correspondance.
    """
    normalized_param = normalize_parameter_name(param_name)
    if not normalized_param:
        logger.warning(f"Paramètre '{param_name}' non normalisé (aucune correspondance)")
        return None
    found_params = []
    # Chercher dans la première colonne (généralement les noms de paramètres)
    for row_idx in range(start_row, min(end_row, ws.max_row + 1)):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value:
            cell_text = str(cell_value).strip()
            found_params.append(cell_text)
            # Correspondance exacte ou partielle
            if normalized_param.lower() in cell_text.lower() or cell_text.lower() in normalized_param.lower():
                logger.info(f"Paramètre '{param_name}' (normalisé: '{normalized_param}') trouvé à la ligne {row_idx} (Excel: '{cell_text}')")
                return row_idx
    logger.warning(f"Paramètre '{param_name}' (normalisé: '{normalized_param}') NON trouvé dans la feuille Excel. Liste Excel: {found_params}")
    return None

def fill_slri_columns_by_index(ws, row_idx, values):
    """
    Remplit les colonnes D, E, F, J, K (index 4, 5, 6, 10, 11) de la ligne row_idx avec les valeurs fournies.
    Args:
        ws: worksheet openpyxl
        row_idx: numéro de ligne à remplir (1-based)
        values: liste [min, max, valeur_mesuree, rejet, mesure_rejet]
    """
    col_indices = [4, 5, 6, 10, 11]  # D, E, F, J, K
    for i, col in enumerate(col_indices):
        if i < len(values):
            val = values[i]
            if val is not None:
                ws.cell(row=row_idx, column=col, value=val)

def find_slri_columns(ws):
    """Trouve les colonnes SLRI dans la feuille Excel.
    
    Recherche les en-têtes suivants:
    - Intervalle acceptable
    - Valeur mesurée de milieu initial
    - Rejet de prj
    - Valeurs Mesure+rejet
    
    Returns:
        dict: {nom_colonne: (index_colonne, ligne_entete)}
    """
    column_mapping = {}
    target_headers = [
        "Intervalle acceptable",
        "Valeur mesurée de milieu initial", 
        "Rejet de prj",
        "Valeurs Mesure+rejet"
    ]
    
    # Chercher dans les 10 premières lignes
    for row_idx in range(1, 11):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_text = str(cell_value).strip()
                for header in target_headers:
                    if header.lower() in cell_text.lower():
                        column_mapping[header] = (col_idx, row_idx)
                        logger.info(f"Colonne '{header}' trouvée à l'index {col_idx}, ligne {row_idx}")
                        break
    
    return column_mapping

def validate_and_color_cells(ws, column_mapping):
    """Valide les seuils et colore les cellules en rouge si dépassement.
    
    Compare 'Valeurs Mesure+rejet' avec 'Intervalle acceptable'.
    Si la valeur dépasse le seuil, colore la cellule en rouge.
    
    Args:
        ws: Feuille de calcul
        column_mapping: Dictionnaire des colonnes SLRI
    """
    if 'Valeurs Mesure+rejet' not in column_mapping or 'Intervalle acceptable' not in column_mapping:
        logger.warning("Colonnes nécessaires pour la validation non trouvées")
        return
    
    mesure_col, header_row = column_mapping['Valeurs Mesure+rejet']
    intervalle_col, _ = column_mapping['Intervalle acceptable']
    
    # Style pour les cellules en dépassement
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    # Parcourir les lignes de données (après les en-têtes)
    for row_idx in range(header_row + 1, ws.max_row + 1):
        try:
            # Récupérer les valeurs
            mesure_cell = ws.cell(row=row_idx, column=mesure_col)
            intervalle_cell = ws.cell(row=row_idx, column=intervalle_col)
            
            mesure_value = mesure_cell.value
            intervalle_value = intervalle_cell.value
            
            # Ignorer si pas de valeur ou si c'est une formule
            if not mesure_value or (isinstance(mesure_value, str) and mesure_value.startswith('=')):
                continue
            
            # Extraire la valeur numérique
            if isinstance(mesure_value, str):
                mesure_num = float(re.sub(r'[^\d.-]', '', mesure_value.split()[0]))
            else:
                mesure_num = float(mesure_value)
            
            # Extraire le seuil de l'intervalle acceptable
            if intervalle_value and isinstance(intervalle_value, str):
                # Formats possibles: "<50", "6-8", ">5", "50", etc.
                if '<' in intervalle_value:
                    seuil = float(re.sub(r'[^\d.-]', '', intervalle_value))
                    if mesure_num >= seuil:
                        mesure_cell.fill = red_fill
                        mesure_cell.font = white_font
                        logger.info(f"Dépassement détecté ligne {row_idx}: {mesure_num} >= {seuil}")
                elif '>' in intervalle_value:
                    seuil = float(re.sub(r'[^\d.-]', '', intervalle_value))
                    if mesure_num <= seuil:
                        mesure_cell.fill = red_fill
                        mesure_cell.font = white_font
                        logger.info(f"Dépassement détecté ligne {row_idx}: {mesure_num} <= {seuil}")
                elif '-' in intervalle_value:
                    # Intervalle min-max
                    parts = intervalle_value.split('-')
                    min_val = float(parts[0].strip())
                    max_val = float(parts[1].strip())
                    if mesure_num < min_val or mesure_num > max_val:
                        mesure_cell.fill = red_fill
                        mesure_cell.font = white_font
                        logger.info(f"Dépassement détecté ligne {row_idx}: {mesure_num} hors [{min_val}, {max_val}]")
        
        except (ValueError, TypeError, IndexError) as e:
            # Ignorer les erreurs de conversion
            continue

def get_sheet_by_phase(wb, phase):
    """Trouve la feuille correspondant à une phase SLRI.
    
    Args:
        wb: Classeur Excel
        phase (str): Phase SLRI ('pre_construction', 'construction', etc.)
        
    Returns:
        Worksheet ou None
    """
    if phase not in SLRI_PHASES:
        return None
    
    possible_names = SLRI_PHASES[phase]
    
    for sheet_name in wb.sheetnames:
        for possible_name in possible_names:
            if possible_name.lower() in sheet_name.lower():
                logger.info(f"Feuille '{sheet_name}' trouvée pour la phase '{phase}'")
                return wb[sheet_name]
    
    return None

def update_slri_excel_macro(file_path: str, data_to_write: dict):
    """
    Met à jour un fichier Excel .xlsm existant sans perdre les macros ni les formules.

    Args:
        file_path (str): Le chemin complet vers le fichier .xlsm.
        data_to_write (dict): Un dictionnaire où les clés sont les adresses des cellules
                              (ex: 'A1', 'Feuille2!B5') et les valeurs sont les données à écrire.

    Returns:
        bool: True si la mise à jour a réussi, False sinon.
    """
    try:
        # Charger le classeur en conservant les macros (VBA)
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)
        logger.info(f"Fichier Excel '{file_path}' chargé avec succès.")

        for cell_address, value in data_to_write.items():
            sheet_name = 'Feuille1'  # Feuille par défaut
            if '!' in cell_address:
                sheet_name, cell_address = cell_address.split('!', 1)
            
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"La feuille '{sheet_name}' n'existe pas. L'écriture pour '{cell_address}' est ignorée.")
                continue

            sheet = workbook[sheet_name]
            sheet[cell_address] = value
            logger.debug(f"Cellule {sheet_name}!{cell_address} mise à jour avec la valeur : {value}")

        # Sauvegarder les modifications dans le même fichier
        workbook.save(file_path)
        logger.info(f"Les modifications ont été sauvegardées dans '{file_path}'.")
        return True

    except FileNotFoundError:
        logger.error(f"Erreur : Le fichier '{file_path}' est introuvable.")
        return False
    except Exception as e:
        logger.error(f"Une erreur inattendue est survenue lors de la mise à jour du fichier Excel : {e}")
        return False

def reorder_dataframe_to_slri_reference(df: pd.DataFrame, slri_reference: pd.DataFrame) -> pd.DataFrame:
    """
    Réordonne le DataFrame fourni selon l'ordre et la structure de la matrice SLRI de référence.
    """
    # On suppose que la colonne 'Paramètre' existe dans les deux DataFrames
    param_order = slri_reference['Paramètre'].tolist()
    milieu_order = slri_reference['Milieu'].tolist() if 'Milieu' in slri_reference.columns else None
    if milieu_order:
        # Réordonne d'abord par milieu puis paramètre
        df['milieu_index'] = df['Milieu'].apply(lambda m: milieu_order.index(m) if m in milieu_order else 999)
        df['param_index'] = df['Paramètre'].apply(lambda p: param_order.index(p) if p in param_order else 999)
        df = df.sort_values(['milieu_index', 'param_index'])
        df = df.drop(columns=['milieu_index', 'param_index'])
    else:
        df['param_index'] = df['Paramètre'].apply(lambda p: param_order.index(p) if p in param_order else 999)
        df = df.sort_values('param_index').drop(columns=['param_index'])
    # Réordonne les colonnes comme dans la matrice SLRI
    ref_cols = [col for col in slri_reference.columns if col in df.columns]
    df = df[ref_cols]
    return df

def extract_min_max_from_intervalle(intervalle):
    if not intervalle:
        return (None, None)
    try:
        if not isinstance(intervalle, str):
            try:
                val = float(intervalle)
                return (None, val)
            except Exception:
                return (None, None)
        s = intervalle.strip().lower()
        s = s.replace(',', '.').replace('–', '-').replace('—', '-').replace('−', '-').replace(' ', '')
        if '-' in s:
            nums = re.findall(r'-?\d+\.?\d*', s)
            if len(nums) >= 2:
                return (float(nums[0]), float(nums[1]))
        if s.startswith('<') or s.startswith('<=') or '≤' in intervalle:
            nums = re.findall(r'\d+\.?\d*', s)
            if nums:
                return (None, float(nums[0]))
        if s.startswith('>') or s.startswith('>=') or '≥' in intervalle:
            nums = re.findall(r'\d+\.?\d*', s)
            if nums:
                return (float(nums[0]), None)
        nums = re.findall(r'\d+\.?\d*', s)
        if nums:
            return (None, float(nums[0]))
        return (None, None)
    except Exception:
        return (None, None)

def parse_numeric(value):
    try:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value)
        nums = re.findall(r'-?\d+\.?\d*', s.replace(',', '.'))
        return float(nums[0]) if nums else None
    except Exception:
        return None

def _normalize_colname(name: str) -> str:
    if not name:
        return ''
    s = remove_accents(str(name)).lower()
    s = re.sub(r'[^a-z0-9]+', '', s)
    return s

def normalize_df_for_slri(df: pd.DataFrame) -> pd.DataFrame:
    """
    Restreint et renomme les colonnes d'un DataFrame vers le schéma SLRI minimal requis.
    Supporte des alias FR/EN et variantes d'écriture.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            'Paramètre', 'Milieu', 'Intervalle acceptable/MIN', 'Intervalle acceptable/MAX',
            'Valeur mesurée de milieux initial', 'Rejet de PHASE CONSTRUCTION',
            'Valeure Mesure+rejet', 'Unité'
        ])

    aliases = {
        # Paramètre
        'parametre': 'Paramètre', 'parametreS': 'Paramètre', 'parametreS': 'Paramètre',
        'parametreparametre': 'Paramètre', 'parameter': 'Paramètre', 'param': 'Paramètre', 'name': 'Paramètre',
        # Milieu
        'milieu': 'Milieu', 'environment': 'Milieu', 'media': 'Milieu', 'environnement': 'Milieu',
        # Intervalle acceptable MIN
        'intervalleacceptablemin': 'Intervalle acceptable/MIN', 'intervallemin': 'Intervalle acceptable/MIN',
        'min': 'Intervalle acceptable/MIN', 'minimum': 'Intervalle acceptable/MIN',
        # Intervalle acceptable MAX
        'intervalleacceptablemax': 'Intervalle acceptable/MAX', 'intervallemax': 'Intervalle acceptable/MAX',
        'max': 'Intervalle acceptable/MAX', 'maximum': 'Intervalle acceptable/MAX',
        # Intervalle acceptable (unique - pour compatibilité)
        'intervalleacceptable': 'Intervalle acceptable', 'reference': 'Intervalle acceptable', 'seuil': 'Intervalle acceptable',
        'acceptableinterval': 'Intervalle acceptable', 'threshold': 'Intervalle acceptable', 'range': 'Intervalle acceptable',
        # Valeur mesurée initiale
        'valeurmesureedemilieuxinitial': 'Valeur mesurée de milieux initial',
        'valeurmesureedemilieuinitial': 'Valeur mesurée de milieux initial',
        'valeurmesuree': 'Valeur mesurée de milieux initial', 'valeur': 'Valeur mesurée de milieux initial',
        'measuredvalue': 'Valeur mesurée de milieux initial', 'value': 'Valeur mesurée de milieux initial',
        # Rejet construction
        'rejetdephaseconstruction': 'Rejet de PHASE CONSTRUCTION', 'rejetdeprj': 'Rejet de PHASE CONSTRUCTION',
        'rejet': 'Rejet de PHASE CONSTRUCTION', 'constructiondischarge': 'Rejet de PHASE CONSTRUCTION',
        # Mesure + rejet
        'valeuremesurerejet': 'Valeure Mesure+rejet', 'valeursmesurerejet': 'Valeure Mesure+rejet',
        'mesurerejet': 'Valeure Mesure+rejet', 'measureplusdischarge': 'Valeure Mesure+rejet', 'total': 'Valeure Mesure+rejet',
        # Unité
        'unite': 'Unité', 'unit': 'Unité', 'units': 'Unité'
    }

    # Construire mapping réel des colonnes
    colmap = {}
    for c in df.columns:
        norm = _normalize_colname(c)
        if norm in aliases:
            target = aliases[norm]
            # ne pas écraser une correspondance déjà fixée
            if target not in colmap:
                colmap[target] = c

    required = ['Paramètre', 'Milieu', 'Intervalle acceptable/MIN', 'Intervalle acceptable/MAX',
                'Valeur mesurée de milieux initial', 'Rejet de PHASE CONSTRUCTION', 'Valeure Mesure+rejet', 'Unité']
    out = pd.DataFrame()
    for key in required:
        if key in colmap:
            out[key] = df[colmap[key]]
        else:
            out[key] = None

    # Calculer Valeure Mesure+rejet si manquante
    if out['Valeure Mesure+rejet'].isnull().all():
        vals = out['Valeur mesurée de milieux initial'].apply(parse_numeric)
        rej = out['Rejet de PHASE CONSTRUCTION'].apply(parse_numeric).fillna(0)
        out['Valeure Mesure+rejet'] = [v + r if v is not None else None for v, r in zip(vals, rej)]

    return out

def update_slri_with_dataframe(file_path: str, df: pd.DataFrame, phase: str = 'pre_construction'):
    """Met à jour un fichier SLRI Excel avec un DataFrame en utilisant le mapping avancé.
    - Mapping paramètre/ligne
    - Coloration automatique
    - Gestion multi-feuilles par phase
    
    Utilise win32com (COM automation) pour préserver parfaitement le format .xlsm
    """
    import shutil
    import os
    
    try:
        # Vérifier si win32com est disponible
        try:
            import win32com.client
            use_win32com = True
        except ImportError:
            logger.warning("win32com non disponible, utilisation d'openpyxl (peut corrompre les .xlsm)")
            use_win32com = False
        
        # Créer une copie de sauvegarde
        backup_path = file_path.replace('.xlsm', '_backup.xlsm').replace('.xlsx', '_backup.xlsx')
        shutil.copy2(file_path, backup_path)
        logger.info(f"Copie de sauvegarde créée: {backup_path}")
        
        if use_win32com:
            return _update_with_win32com(file_path, df, phase, backup_path)
        else:
            return _update_with_openpyxl(file_path, df, phase)

    except Exception as e:
        import traceback
        logger.error(f"Erreur lors de la mise à jour du fichier SLRI: {e}")
        logger.error(traceback.format_exc())
        return False

def _update_with_win32com(file_path: str, df: pd.DataFrame, phase: str, backup_path: str):
    """Mise à jour via win32com (COM automation Excel) - méthode la plus sûre pour .xlsm"""
    import win32com.client
    import os
    
    excel = None
    wb = None
    try:
        # Normaliser le DataFrame
        df = normalize_df_for_slri(df)
        
        # Démarrer Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Ouvrir le fichier
        abs_path = os.path.abspath(file_path)
        wb = excel.Workbooks.Open(abs_path)
        
        # Trouver la feuille correspondante à la phase
        ws = None
        phase_names = SLRI_PHASES.get(phase.lower().replace(' ', '_').replace('é', 'e').replace('è', 'e'), [])
        for sheet in wb.Sheets:
            if sheet.Name in phase_names:
                ws = sheet
                break
        if not ws:
            ws = wb.Sheets(1)
        
        logger.info(f"Feuille sélectionnée: {ws.Name}")
        
        # Trouver les colonnes SLRI par en-tête (recherche dynamique)
        col_d = None  # Intervalle acceptable MIN
        col_e = None  # Intervalle acceptable MAX
        col_f = None  # Valeur mesurée initiale
        col_j = None  # Rejet de PHASE CONSTRUCTION
        col_k = None  # Valeure Mesure+rejet
        
        # Construire une carte de toutes les cellules pour gérer les fusions
        header_map = {}
        for row_idx in range(1, 11):
            for col_idx in range(1, min(ws.UsedRange.Columns.Count + 1, 30)):
                try:
                    cell_val = ws.Cells(row_idx, col_idx).Value
                    if cell_val:
                        cell_str = str(cell_val).strip().lower()
                        if cell_str:
                            if col_idx not in header_map:
                                header_map[col_idx] = []
                            header_map[col_idx].append(cell_str)
                except:
                    continue
        
        # Analyser les en-têtes pour identifier les colonnes
        intervalle_cols = []
        for col_idx, headers in sorted(header_map.items()):
            combined = ' '.join(headers)
            
            # Colonnes D et E : Intervalle acceptable (deux colonnes consécutives)
            if 'intervalle' in combined and 'acceptable' in combined:
                intervalle_cols.append(col_idx)
            
            # Colonne F : Valeur mesurée
            if not col_f and 'valeur' in combined and 'mesur' in combined and 'rejet' not in combined:
                col_f = col_idx
            
            # Colonne J : Rejet
            if not col_j and 'rejet' in combined and ('phase' in combined or 'prj' in combined or 'construction' in combined):
                col_j = col_idx
            
            # Colonne K : Valeur Mesure+rejet
            if not col_k and 'valeur' in combined and 'mesure' in combined and 'rejet' in combined:
                col_k = col_idx
        
        # Assigner D et E depuis les colonnes d'intervalle trouvées
        if len(intervalle_cols) >= 2:
            col_d = intervalle_cols[0]
            col_e = intervalle_cols[1]
        elif len(intervalle_cols) == 1:
            col_d = intervalle_cols[0]
            col_e = intervalle_cols[0] + 1
        
        logger.info(f"Colonnes 'Intervalle acceptable' trouvées: {intervalle_cols}")
        logger.info(f"En-têtes détectés: {dict(list(header_map.items())[:15])}")
        
        # Si colonnes non trouvées par en-tête, utiliser les index par défaut
        if not col_d: col_d = 4
        if not col_e: col_e = 5
        if not col_f: col_f = 6
        if not col_j: col_j = 10
        if not col_k: col_k = 11
        
        logger.info(f"Colonnes SLRI identifiées: D={col_d}, E={col_e}, F={col_f}, J={col_j}, K={col_k}")

        # Parcourir le DataFrame et remplir les cellules
        logger.info(f"Début du traitement de {len(df)} paramètres")
        
        for idx, row in df.iterrows():
            param_name = row.get('Paramètre') or row.get('Parametre')
            if not param_name:
                continue
            
            # Parser les valeurs MIN/MAX
            min_val = row.get('Intervalle acceptable/MIN')
            max_val = row.get('Intervalle acceptable/MAX')
            
            # Fallback : si colonnes séparées non disponibles, essayer de parser une colonne unique
            if min_val is None and max_val is None:
                intervalle = (row.get('Intervalle acceptable') or row.get('Reference') or '')
                min_val, max_val = extract_min_max_from_intervalle(intervalle)
            
            value = (row.get('Valeur mesurée de milieux initial') or 
                     row.get('Valeur mesurée') or 
                     row.get('Valeur') or None)
            v_num = parse_numeric(value)
            
            rejet = (row.get('Rejet de PHASE CONSTRUCTION') or row.get('Rejet de prj') or row.get('Rejet') or 0)
            r_num = parse_numeric(rejet)
            if r_num is None:
                r_num = 0
            
            mesure_rejet = (v_num + r_num) if v_num is not None else None
            
            # Trouver la ligne dans Excel (recherche dans colonnes A, B, C)
            # IMPORTANT : Commencer à la ligne 2 (après la ligne d'en-tête principale)
            target_row = None
            param_lower = param_name.strip().lower()
            # Normaliser : supprimer accents, espaces, caractères spéciaux
            try:
                param_normalized = remove_accents(param_lower).replace(' ', '').replace('(', '').replace(')', '').replace('-', '').replace('_', '')
            except Exception as e:
                logger.error(f"Erreur de normalisation pour '{param_name}': {e}")
                param_normalized = param_lower
            
            for i in range(2, min(ws.UsedRange.Rows.Count + 1, 300)):  # Commencer à la ligne 2 (après la ligne d'en-tête principale)
                for col in [1, 2, 3]:  # Colonnes A, B, C
                    try:
                        cell_val = ws.Cells(i, col).Value
                        if cell_val:
                            cell_str = str(cell_val).strip().lower()
                            # Éviter les en-têtes (mots-clés communs)
                            if any(keyword in cell_str for keyword in ['intervalle', 'valeur', 'rejet', 'score', 'milieu', 'paramètre', 'unité']):
                                continue
                            
                            # Normaliser la cellule Excel
                            cell_normalized = remove_accents(cell_str).replace(' ', '').replace('(', '').replace(')', '').replace('-', '').replace('_', '')
                            
                            # Correspondance exacte normalisée
                            if param_normalized == cell_normalized:
                                target_row = i
                                break
                            # Correspondance partielle
                            elif param_normalized in cell_normalized or cell_normalized in param_normalized:
                                # Vérifier que c'est une vraie correspondance (au moins 3 caractères)
                                if len(param_normalized) >= 3 and len(cell_normalized) >= 3:
                                    target_row = i
                                    break
                    except:
                        continue
                if target_row:
                    break
            
            if not target_row:
                logger.warning(f"Paramètre '{param_name}' (normalisé: '{param_normalized}') non trouvé dans la feuille Excel")
                continue
            
            logger.info(f"Paramètre '{param_name}' trouvé à la ligne {target_row}")
            
            # Remplir toutes les colonnes SLRI : D, E, F, J, K
            
            # Colonnes D et E : Intervalle acceptable MIN/MAX
            if min_val is not None:
                ws.Cells(target_row, col_d).Value = min_val
            if max_val is not None:
                ws.Cells(target_row, col_e).Value = max_val
            
            # Colonne F : Valeur mesurée initiale
            if v_num is not None:
                ws.Cells(target_row, col_f).Value = v_num
            
            # Colonne J : Rejet de PHASE CONSTRUCTION
            if r_num is not None:
                ws.Cells(target_row, col_j).Value = r_num
            
            # Colonne K : Valeure Mesure+rejet (ne pas écraser si formule)
            k_cell_formula = ws.Cells(target_row, col_k).Formula
            if not k_cell_formula or not str(k_cell_formula).startswith('='):
                if mesure_rejet is not None:
                    ws.Cells(target_row, col_k).Value = mesure_rejet

        # Sauvegarder et fermer
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        
        logger.info(f"Fichier SLRI mis à jour avec succès via win32com: {file_path}")
        return True
        
    except Exception as e:
        import traceback
        logger.error(f"Erreur win32com: {e}")
        logger.error(traceback.format_exc())
        return False
    finally:
        # Nettoyer les objets COM
        try:
            if wb:
                wb.Close(SaveChanges=False)
            if excel:
                excel.Quit()
        except:
            pass

def _update_with_openpyxl(file_path: str, df: pd.DataFrame, phase: str):
    """Fallback avec openpyxl (peut corrompre les .xlsm)"""
    wb = None
    try:
        df = normalize_df_for_slri(df)
        wb = openpyxl.load_workbook(file_path, keep_vba=True, data_only=False, keep_links=True)
        ws = get_sheet_by_phase(wb, phase)
        if not ws:
            ws = wb.active
        
        for _, row in df.iterrows():
            param_name = row.get('Paramètre') or row.get('Parametre')
            if not param_name:
                continue
            
            intervalle = (row.get('Intervalle acceptable') or row.get('Reference') or '')
            min_val, max_val = extract_min_max_from_intervalle(intervalle)
            value = (row.get('Valeur mesurée de milieux initial') or row.get('Valeur mesurée') or row.get('Valeur') or None)
            v_num = parse_numeric(value)
            rejet = (row.get('Rejet de PHASE CONSTRUCTION') or row.get('Rejet de prj') or row.get('Rejet') or 0)
            r_num = parse_numeric(rejet)
            if r_num is None:
                r_num = 0
            mesure_rejet = (v_num + r_num) if v_num is not None else None
            
            target_row = find_parameter_row(ws, param_name)
            if not target_row:
                logger.warning(f"Paramètre '{param_name}' non trouvé")
                continue
            
            k_cell = ws.cell(row=target_row, column=11)
            k_val = None if (isinstance(k_cell.value, str) and k_cell.value.startswith('=')) else mesure_rejet
            v_out = v_num if v_num is not None else value
            fill_slri_columns_by_index(ws, target_row, [min_val, max_val, v_out, r_num, k_val])
        
        column_mapping = find_slri_columns(ws)
        if column_mapping:
            validate_and_color_cells(ws, column_mapping)
        
        import os
        base_name = os.path.splitext(file_path)[0]
        ext = os.path.splitext(file_path)[1]
        output_path = f"{base_name}_updated{ext}"
        wb.save(output_path)
        
        try:
            os.remove(file_path)
            os.rename(output_path, file_path)
        except:
            return output_path
        
        return True
    except Exception as e:
        import traceback
        logger.error(traceback.format_exc())
        return False
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass

# --- Exemple d'utilisation (peut être retiré ou commenté) ---
if __name__ == '__main__':
    # Remplacer par le chemin réel de votre fichier
    target_file = 'E:\\Users\\tahaa\\OneDrive\\Bureau\\1test risque\\Nouveau dossier (3)\\Macro Standardiser-Levaluation-des-risques-et-impacts 012.xlsm'
    
    # Données à écrire. Format: {"Feuille!Cellule": valeur}
    # Si la feuille n'est pas spécifiée, la première feuille est utilisée par défaut.
    data = {
        'Feuil1!B5': 42,               # Mettre à jour B5 sur 'Feuil1'
        'Feuil1!C10': 'Test de mise à jour', # Mettre à jour C10 sur 'Feuil1'
        'AutreFeuille!A1': 999        # Mettre à jour A1 sur 'AutreFeuille'
    }
    
    # Appeler la fonction de mise à jour
    success = update_slri_excel_macro(target_file, data)
    
    if success:
        print("Le fichier Excel a été mis à jour avec succès.")
    else:
        print("Échec de la mise à jour du fichier Excel.")
