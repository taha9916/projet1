"""
Module d'intégration SLRI (Standardiser l'évaluation des risques et impacts)
Intègre la méthodologie SLRI dans le système d'analyse environnementale
"""

import os
import json
import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

logger = logging.getLogger(__name__)

class SLRIAnalyzer:
    """
    Analyseur SLRI pour l'évaluation standardisée des risques environnementaux
    """
    
    def __init__(self, slri_path: str = None):
        """
        Initialise l'analyseur SLRI
        
        Args:
            slri_path: Chemin vers le dossier SLRI
        """
        self.slri_path = Path(slri_path) if slri_path else Path("SLRI")
        self.phases = ["PRE CONSTRUCTION", "CONSTRUCTION", "exploitation", "démantalement"]
        self.milieux = {
            "PHYSIQUE/EAU": ["Température", "Ph", "Turbidité", "Conductivité", "DBO5", "DCO", 
                           "Oxygène dissous", "Nitrates", "Nitrites", "Ammoniac", "Phosphore total", 
                           "Azote total", "Métaux lourds", "Hydrocarbures"],
            "PHYSIQUE/SOL": ["pH", "Perméabilité", "Matière organique", "Carbone organique", 
                           "Métaux lourds", "Azote total", "Phosphore total"],
            "PHYSIQUE/AIR": ["Poussières totales", "PM10", "PM2.5", "SO₂", "NOx", "CO", "O₃"],
            "BIOLOGIQUE/FLORE": ["Biodiversité terrestre", "Biodiversité marine"],
            "BIOLOGIQUE/FAUNE": ["mammifères", "amphibiens", "reptiles"]
        }
        
        # Échelles d'évaluation SLRI
        self.echelles = {
            "score_parametre": {0: "Conforme", 1: "Dépassement léger (≤10%)", 2: "Dépassement important (>10%)"},
            "duree": {0: "Très court terme", 1: "Court terme", 2: "Moyen terme", 3: "Long terme", 4: "Permanent"},
            "etendue": {0: "Locale", 1: "Régionale", 2: "Nationale", 3: "Internationale"},
            "frequence": {0: "Très rare", 1: "Rare", 2: "Occasionnelle", 3: "Fréquente", 4: "Permanente"},
            "amplitude": {
                "0-4": "Faible", "5-8": "Moyen", "9-12": "Fort", "13+": "Très grave"
            }
        }
        
        self.logger = logging.getLogger(__name__)
    
    def load_slri_data(self):
        """
        Charge les données SLRI depuis le fichier Excel de référence
        """
        try:
            excel_path = os.path.join(self.slri_path, "Standardiser-levaluation-des-risques-et-impacts.xlsx")
            
            if not os.path.exists(excel_path):
                logger.warning(f"Fichier Excel SLRI non trouvé: {excel_path}")
                return self._load_from_text_files()  # Fallback vers les fichiers texte
            
            # Charger le fichier Excel de référence
            wb = load_workbook(excel_path)
            
            # Charger la matrice principale
            ws_matrice = wb["Matrice SLRI"]
            matrice_data = []
            
            # Lire les données de la matrice (à partir de la ligne 2)
            for row in ws_matrice.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si la phase n'est pas vide
                    matrice_data.append({
                        "Phase": row[0],
                        "Milieu": row[1],
                        "Paramètre": row[2],
                        "Valeur_Mesuree": row[3],
                        "Unité": row[4],
                        "Valeur_Reference": row[5],
                        "Score_Parametre": row[6],
                        "Duree": row[7],
                        "Etendue": row[8],
                        "Frequence": row[9],
                        "Amplitude": row[10],
                        "Classification": row[11]
                    })
            
            # Charger les échelles
            ws_echelles = wb["Échelles"]
            echelles_data = []
            for row in ws_echelles.iter_rows(values_only=True):
                if row[0]:
                    echelles_data.append(row)
            
            slri_data = {
                "matrice": matrice_data,
                "echelles": echelles_data,
                "structure_excel": True
            }
            
            logger.info(f"Données SLRI chargées depuis Excel: {len(matrice_data)} paramètres")
            return slri_data
            
        except Exception as e:
            logger.error(f"Erreur lors du chargement du fichier Excel SLRI: {e}")
            return self._load_from_text_files()  # Fallback
    
    def _load_from_text_files(self):
        """
        Méthode de fallback pour charger depuis les fichiers texte
        """
        slri_data = {}
        
        try:
            # Charger les données de chaque phase
            phase_files = {
                "PRE CONSTRUCTION": "PRE CONSTRUCTION.txt",
                "CONSTRUCTION": "CONSTRUCTION.txt", 
                "exploitation": "exploitation.txt",
                "démantalement": "démantalement.txt"
            }
            
            for phase_name, filename in phase_files.items():
                file_path = os.path.join(self.slri_path, filename)
                if os.path.exists(file_path):
                    phase_data = self._parse_phase_file(file_path)
                    slri_data[phase_name] = phase_data
                else:
                    logger.warning(f"Fichier SLRI non trouvé: {file_path}")
            
            # Charger les échelles et la matrice d'impacts
            echelles_path = os.path.join(self.slri_path, "Echelles.txt")
            if os.path.exists(echelles_path):
                slri_data["echelles"] = self._parse_echelles_file(echelles_path)
            
            matrice_path = os.path.join(self.slri_path, "matrice d'impacts.txt")
            if os.path.exists(matrice_path):
                slri_data["matrice_impacts"] = self._parse_matrice_file(matrice_path)
            
            logger.info(f"Données SLRI chargées depuis fichiers texte: {len(slri_data)} éléments")
            return slri_data
            
        except Exception as e:
            logger.error(f"Erreur lors du chargement des données SLRI: {e}")
            return {}
    
    def _parse_phase_file(self, file_path: str) -> dict:
        """
        Parse un fichier de phase SLRI
        
        Args:
            file_path: Chemin vers le fichier
            
        Returns:
            Dict contenant les données parsées
        """
        phase_data = {
            "parametres": [],
            "scores": [],
            "evaluations": []
        }
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            for line in lines[1:]:  # Skip header
                if line.strip() and not line.startswith('\t'):
                    parts = line.split('\t')
                    if len(parts) >= 5:
                        param_data = {
                            "milieu": parts[0].strip(),
                            "parametre": parts[1].strip(),
                            "unite": parts[2].strip(),
                            "intervalle": parts[3].strip(),
                            "valeur_mesuree": parts[4].strip()
                        }
                        phase_data["parametres"].append(param_data)
            
            return phase_data
            
        except Exception as e:
            self.logger.error(f"Erreur lors du parsing de {file_path}: {e}")
            return phase_data
    
    def _parse_matrice_impacts(self, file_path: str) -> dict:
        """
        Parse la matrice d'impacts
        
        Args:
            file_path: Chemin vers le fichier matrice
            
        Returns:
            Dict contenant la matrice d'impacts
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Structure de base de la matrice
            matrice = {
                "milieux": ["Milieu physique", "Milieu biologique", "Milieu humain"],
                "composantes": {
                    "Milieu physique": ["sol", "air", "Qualité des eaux", "Paysage"],
                    "Milieu biologique": ["Flore", "faune"],
                    "Milieu humain": ["Population et habitats", "Structures et moyens de subsistance", 
                                    "Santé & Sécurité", "Ambiance sonore", "Activité socio-économique/emplois",
                                    "Infrastructures et circulation routière"]
                },
                "sources_impact": {
                    "pre_construction": ["Prospection préliminaire", "Transport et circulation", "Installation du chantier"],
                    "construction": ["Transport et circulation", "Excavations", "Présence de la Base"],
                    "exploitation": ["Bâtiments et équipements", "Rejets liquides et solides", "Présence des infrastructures"],
                    "demantelement": ["Remise en état", "Mise en œuvre du projet", "Entretien et réparation", "Excavations", "Démantèlement"]
                }
            }
            
            return matrice
            
        except Exception as e:
            self.logger.error(f"Erreur lors du parsing de la matrice: {e}")
            return {}
    
    def calculate_risk_score(self, parametre_score: int, duree: int, etendue: int, frequence: int) -> dict:
        """
        Calcule le score de risque selon la méthodologie SLRI
        
        Args:
            parametre_score: Score du paramètre (0-2)
            duree: Score de durée (0-4)
            etendue: Score d'étendue (0-3)
            frequence: Score de fréquence (0-4)
            
        Returns:
            Dict avec le score total et l'amplitude
        """
        # Calcul du score total selon SLRI: Durée × Étendue × Fréquence
        score_temporel_spatial = duree * etendue * frequence
        
        # Score final combiné
        score_final = parametre_score + score_temporel_spatial
        
        # Détermination de l'amplitude
        if score_final <= 4:
            amplitude = "Faible"
        elif score_final <= 8:
            amplitude = "Moyen"
        elif score_final <= 12:
            amplitude = "Fort"
        else:
            amplitude = "Très grave"
        
        return {
            "score_parametre": parametre_score,
            "score_temporel_spatial": score_temporel_spatial,
            "score_final": score_final,
            "amplitude": amplitude,
            "details": {
                "duree": self.echelles["duree"][duree],
                "etendue": self.echelles["etendue"][etendue],
                "frequence": self.echelles["frequence"][frequence]
            }
        }
    
    def evaluate_environmental_parameter(self, valeur_mesuree: float, intervalle_acceptable: str, 
                                       duree: int = 1, etendue: int = 1, frequence: int = 1) -> dict:
        """
        Évalue un paramètre environnemental selon SLRI
        
        Args:
            valeur_mesuree: Valeur mesurée du paramètre
            intervalle_acceptable: Intervalle acceptable (ex: "10-25", "≤50", ">1")
            duree: Score de durée (0-4)
            etendue: Score d'étendue (0-3)
            frequence: Score de fréquence (0-4)
            
        Returns:
            Dict avec l'évaluation complète
        """
        # Calcul du score du paramètre
        parametre_score = self._calculate_parameter_score(valeur_mesuree, intervalle_acceptable)
        
        # Calcul du score de risque total
        risk_assessment = self.calculate_risk_score(parametre_score, duree, etendue, frequence)
        
        return {
            "valeur_mesuree": valeur_mesuree,
            "intervalle_acceptable": intervalle_acceptable,
            "conforme": parametre_score == 0,
            "evaluation": risk_assessment
        }
    
    def _calculate_parameter_score(self, valeur: float, intervalle: str) -> int:
        """
        Calcule le score d'un paramètre selon son intervalle acceptable
        
        Args:
            valeur: Valeur mesurée
            intervalle: Intervalle acceptable
            
        Returns:
            int: Score du paramètre (0=conforme, 1=attention, 2=non-conforme)
        """
        try:
            if isinstance(valeur, str) and valeur == "":
                return ""
            
            val = float(valeur)
            
            # Traitement des intervalles
            if "<" in str(intervalle):
                seuil = float(str(intervalle).replace("<", "").replace("°C", "").replace("mg/L", "").replace("µg/m³", "").replace("NTU", "").replace("µS/cm", "").strip())
                return 0 if val <= seuil else 2
            elif ">" in str(intervalle):
                seuil = float(str(intervalle).replace(">", "").replace("%", "").replace("mg/L", "").strip())
                return 0 if val >= seuil else 2
            else:
                return 0
        
        except (ValueError, TypeError):
            return ""

    def analyze_with_complete_slri_structure(self, coordinates, project_type="industriel"):
        """
        Analyse complète selon la structure SLRI de référence avec tous les paramètres
        """
        try:
            # Charger la structure SLRI de référence
            slri_data = self.load_slri_data()
            
            # Simuler la collecte de données environnementales
            collected_data = self._simulate_data_collection(coordinates)
            
            # Créer la structure complète SLRI
            complete_slri_analysis = self._create_complete_slri_structure(collected_data, coordinates, project_type)
            
            logger.info(f"Analyse SLRI complète générée pour les coordonnées {coordinates}")
            return complete_slri_analysis
            
        except Exception as e:
            logger.error(f"Erreur lors de l'analyse SLRI complète: {e}")
            return None

    def _create_complete_slri_structure(self, collected_data, coordinates, project_type):
        """
        Crée la structure SLRI complète selon les tableaux de référence
        """
        # Structure des paramètres SLRI de référence
        slri_reference_structure = {
            "eau": {
            'Température': '<30°C',
            'pH': '6-8',
            'Turbidité': '<5 NTU',
            'Conductivité': '<1000 µS/cm',
            'DBO5': '<5 mg/L',
            'DCO': '<25 mg/L',
            'Oxygène dissous': '>5 mg/L',
            'Nitrates': '<50 mg/L',
            'Nitrites': '<0.5 mg/L',
            'Ammoniac': '<0.5 mg/L',
            'Phosphore total': '<0.1 mg/L',
            'Azote total': '<10 mg/L',
            'Plomb (Pb)': '<0.01 mg/L',
            'Cadmium (Cd)': '<0.005 mg/L',
            'Chrome (Cr)': '<0.05 mg/L',
            'Cuivre (Cu)': '<2 mg/L',
            'Zinc (Zn)': '<3 mg/L',
            'Nickel (Ni)': '<0.07 mg/L',
            'Mercure (Hg)': '<0.001 mg/L',
            'Arsenic (As)': '<0.01 mg/L',
            'Hydrocarbures totaux (HCT)': '<0.05 mg/L',
            'Hydrocarbures aromatiques polycycliques (HAP)': '<0.0002 mg/L'
        },
        "sol": {
            'pH': '6-8',
            'Perméabilité': '10-6 à 10-4 m/s',
            'Matière organique': '2-5%',
            'Carbone organique': '1-3%',
            'Plomb (Pb)': '<85 mg/kg',
            'Cadmium (Cd)': '<1.4 mg/kg',
            'Chrome (Cr)': '<100 mg/kg',
            'Cuivre (Cu)': '<36 mg/kg',
            'Zinc (Zn)': '<140 mg/kg',
            'Nickel (Ni)': '<35 mg/kg',
            'Mercure (Hg)': '<0.4 mg/kg',
            'Arsenic (As)': '<12 mg/kg',
            'Azote total': '0.1-0.5%',
            'Phosphore total': '400-1200 mg/kg'
        },
        "air": {
            'Poussières totales': '<150 µg/m³',
            'PM10': '<50 µg/m³',
            'PM2.5': '<25 µg/m³',
            'SO2': '<125 µg/m³',
            'NOx': '<200 µg/m³',
            'CO': '<10 mg/m³',
            'O3 (ozone)': '<120 µg/m³'
        },
        "biologique": {
            'Flore terrestre': 'Présence/Absence',
            'Flore marine': 'Présence/Absence',
            'Mammifères': 'Présence/Absence',
            'Amphibiens': 'Présence/Absence',
            'Reptiles': 'Présence/Absence',
            'Statut de protection': 'Oui/Non',
            'Présence sur site': 'Oui/Non'
        },
        "humain": {
            'Population riveraine': 'Nombre d\'habitants',
            'Distance habitations': 'mètres',
            'Activités économiques': 'Type',
            'Patrimoine culturel': 'Présence/Absence',
            'Accès aux ressources': 'Oui/Non'
        }
    }
    
    # Créer la structure complète pour la phase PRE CONSTRUCTION
    phase_parametres = {}
    major_risks = []
    
    for milieu, parametres in slri_reference_structure.items():
        phase_parametres[milieu] = {}
        
        for param_name, seuil in parametres.items():
            # Récupérer la valeur mesurée si disponible
            valeur_mesuree = ""
            score = ""
            amplitude = "NON ANALYSÉ"
            
            if milieu.upper() in collected_data and param_name in collected_data[milieu.upper()]:
                valeur_data = collected_data[milieu.upper()][param_name]
                if isinstance(valeur_data, tuple):
                    valeur_mesuree = f"{valeur_data[0]} {valeur_data[1]}".strip()
                else:
                    valeur_mesuree = str(valeur_data)
                
                # Calculer le score si la valeur est numérique
                if valeur_data[0] and isinstance(valeur_data[0], (int, float)):
                    score = self._score_parameter(valeur_data[0], seuil)
                    if score == 0:
                        amplitude = "CONFORME"
                    elif score == 1:
                        amplitude = "MOYEN"
                    else:
                        amplitude = "FORT"
                        
                    # Ajouter aux risques majeurs si nécessaire
                    if score >= 1:
                        major_risks.append({
                            'parameter': param_name,
                            'milieu': milieu.upper(),
                            'amplitude': amplitude,
                            'valeur': valeur_mesuree,
                            'seuil': seuil
                        })
    
    def _generate_phase_report(self, phase, parametres_data):
        """
        Génère un rapport pour une phase spécifique
        
        Args:
            phase: Nom de la phase
            parametres_data: Liste des données de paramètres
            
        Returns:
            Dict contenant le rapport de phase
        """
        evaluations = []
        scores_totaux = []
        
        for param_data in parametres_data:
            try:
                valeur = float(param_data.get("valeur_mesuree", 0))
                intervalle = param_data.get("intervalle", "")
                
                evaluation = self.evaluate_environmental_parameter(
                    valeur, intervalle,
                    duree=param_data.get("duree", 1),
                    etendue=param_data.get("etendue", 1),
                    frequence=param_data.get("frequence", 1)
                )
                
                evaluation["parametre"] = param_data.get("parametre", "")
                evaluation["milieu"] = param_data.get("milieu", "")
                evaluations.append(evaluation)
                scores_totaux.append(evaluation["evaluation"]["score_final"])
                
            except (ValueError, KeyError) as e:
                self.logger.warning(f"Erreur lors de l'évaluation du paramètre {param_data}: {e}")
        
        # Calcul des statistiques de phase
        if scores_totaux:
            score_moyen = np.mean(scores_totaux)
            score_max = max(scores_totaux)
            nb_non_conformes = sum(1 for eval in evaluations if not eval["conforme"])
        else:
            score_moyen = score_max = nb_non_conformes = 0
        
        return {
            "phase": phase,
            "evaluations": evaluations,
            "statistiques": {
                "score_moyen": score_moyen,
                "score_maximum": score_max,
                "nb_parametres": len(evaluations),
                "nb_non_conformes": nb_non_conformes,
                "taux_conformite": (len(evaluations) - nb_non_conformes) / len(evaluations) * 100 if evaluations else 0
            }
        }
    
    def generate_global_assessment(self, slri_data: dict) -> dict:
        """
        Génère une évaluation globale SLRI
        
        Args:
            slri_data: Données SLRI complètes
            
        Returns:
            Dict contenant l'évaluation globale
        """
        phase_reports = {}
        global_stats = {
            "scores_par_phase": {},
            "risques_majeurs": [],
            "recommandations": []
        }
        
        # Analyse par phase
        for phase in self.phases:
            if phase in slri_data:
                phase_data = slri_data[phase]
                report = self.generate_phase_report(phase, phase_data.get("parametres", []))
                phase_reports[phase] = report
                global_stats["scores_par_phase"][phase] = report["statistiques"]["score_moyen"]
                
                # Identification des risques majeurs
                for eval in report["evaluations"]:
                    if eval["evaluation"]["amplitude"] in ["Fort", "Très grave"]:
                        global_stats["risques_majeurs"].append({
                            "phase": phase,
                            "parametre": eval["parametre"],
                            "milieu": eval["milieu"],
                            "amplitude": eval["evaluation"]["amplitude"],
                            "score": eval["evaluation"]["score_final"]
                        })
        
        # Génération de recommandations
        global_stats["recommandations"] = self._generate_recommendations(global_stats["risques_majeurs"])
        
        return {
            "rapports_phases": phase_reports,
            "statistiques_globales": global_stats,
            "date_evaluation": pd.Timestamp.now().isoformat()
        }
    
    def _generate_recommendations(self, risques_majeurs: list) -> list:
        """
        Génère des recommandations basées sur les risques majeurs identifiés
        
        Args:
            risques_majeurs: Liste des risques majeurs
            
        Returns:
            Liste de recommandations
        """
        recommendations = []
        
        # Analyse par milieu
        milieux_impactes = {}
        for risque in risques_majeurs:
            milieu = risque["milieu"]
            if milieu not in milieux_impactes:
                milieux_impactes[milieu] = []
            milieux_impactes[milieu].append(risque)
        
        # Recommandations spécifiques par milieu
        for milieu, risques in milieux_impactes.items():
            if "EAU" in milieu:
                recommendations.append(f"Mettre en place un système de traitement des eaux pour le milieu {milieu}")
                recommendations.append("Installer des dispositifs de surveillance continue de la qualité de l'eau")
            elif "SOL" in milieu:
                recommendations.append(f"Prévoir des mesures de protection et de réhabilitation des sols")
                recommendations.append("Mettre en place un plan de gestion des terres excavées")
            elif "AIR" in milieu:
                recommendations.append("Installer des systèmes de dépoussiérage et de filtration")
                recommendations.append("Mettre en place un plan de surveillance de la qualité de l'air")
        
        # Recommandations générales
        if len(risques_majeurs) > 5:
            recommendations.append("Réviser la conception du projet pour réduire les impacts environnementaux")
        
        recommendations.append("Élaborer un plan de gestion environnementale détaillé")
        recommendations.append("Mettre en place un système de monitoring environnemental continu")
        
        return recommendations
    
    def _simulate_data_collection(self, coordinates):
        """
        Simule la collecte de données environnementales pour les coordonnées données
        """
        lat, lon = coordinates
        
        # Simulation de données environnementales typiques
        simulated_data = {
            "EAU": {
                "Température": (22.5, "°C"),
                "pH": (7.2, ""),
                "Turbidité": (3.2, "NTU"),
                "Conductivité": (850, "µS/cm"),
                "DBO5": (4.1, "mg/L"),
                "Oxygène dissous": (6.8, "mg/L"),
                "Nitrates": (35, "mg/L"),
                "Phosphore total": (0.08, "mg/L")
            },
            "SOL": {
                "pH": (7.1, ""),
                "Matière organique": (3.2, "%"),
                "Carbone organique": (1.8, "%"),
                "Plomb (Pb)": (45, "mg/kg"),
                "Cadmium (Cd)": (0.8, "mg/kg")
            },
            "AIR": {
                "PM10": (42, "µg/m³"),
                "PM2.5": (28, "µg/m³"),
                "SO2": (95, "µg/m³"),
                "NOx": (180, "µg/m³")
            },
            "BIOLOGIQUE": {
                "Flore terrestre": ("Présence", ""),
                "Mammifères": ("Présence", ""),
                "Statut protection": ("Non", "")
            },
            "HUMAIN": {
                "Population riveraine": (2500, "habitants"),
                "Distance habitations": (800, "m"),
                "Patrimoine culturel": ("Absence", "")
            }
        }
        
        return simulated_data
    
    def _map_data_to_slri_structure(self, collected_data, slri_data):
        """
        Mappe les données collectées sur la structure SLRI de référence
        """
        mapped_data = []
        
        if "matrice" in slri_data:
            for matrix_entry in slri_data["matrice"]:
                phase = matrix_entry["Phase"]
                milieu = matrix_entry["Milieu"]
                parametre = matrix_entry["Paramètre"]
                reference = matrix_entry["Valeur_Reference"]
                
                # Chercher la valeur correspondante dans les données collectées
                valeur_mesuree = ""
                unite = matrix_entry["Unité"]
                
                if milieu in collected_data and parametre in collected_data[milieu]:
                    valeur_mesuree, unite_collectee = collected_data[milieu][parametre]
                    if unite_collectee:
                        unite = unite_collectee
                
                mapped_entry = {
                    "Phase": phase,
                    "Milieu": milieu,
                    "Paramètre": parametre,
                    "Valeur_Mesuree": valeur_mesuree,
                    "Unité": unite,
                    "Valeur_Reference": reference,
                    "Score_Parametre": "",
                    "Duree": "",
                    "Etendue": "",
                    "Frequence": "",
                    "Amplitude": "",
                    "Classification": ""
                }
                
                mapped_data.append(mapped_entry)
        
        return mapped_data
    
    def _calculate_slri_scores(self, mapped_data):
        """
        Calcule les scores SLRI pour les données mappées
        """
        for entry in mapped_data:
            if entry["Valeur_Mesuree"] and entry["Valeur_Reference"]:
                # Calculer le score du paramètre
                score_param = self._score_parameter(entry["Valeur_Mesuree"], entry["Valeur_Reference"])
                entry["Score_Parametre"] = score_param
                
                # Assigner des valeurs par défaut pour durée, étendue, fréquence
                entry["Duree"] = 2  # Moyenne durée par défaut
                entry["Etendue"] = 1  # Locale par défaut
                entry["Frequence"] = 2  # Intermittente par défaut
                
                # Calculer l'amplitude
                amplitude = score_param * (entry["Duree"] + entry["Etendue"] + entry["Frequence"])
                entry["Amplitude"] = amplitude
                
                # Classification
                if amplitude <= 4:
                    entry["Classification"] = "FAIBLE"
                elif amplitude <= 8:
                    entry["Classification"] = "MOYEN"
                elif amplitude <= 12:
                    entry["Classification"] = "FORT"
                else:
                    entry["Classification"] = "TRÈS GRAVE"
        
        return mapped_data
    
    def _generate_slri_assessment(self, scored_data, coordinates):
        """
        Génère l'évaluation SLRI selon le format de référence
        """
        # Organiser les données par phase
        rapports_phases = {}
        phases = ["PRE CONSTRUCTION", "CONSTRUCTION", "EXPLOITATION", "DÉMANTÈLEMENT"]
        
        for phase in phases:
            phase_data = [entry for entry in scored_data if entry["Phase"] == phase]
            
            parametres = []
            for entry in phase_data:
                param_dict = {
                    "milieu": entry["Milieu"],
                    "parametre": entry["Paramètre"],
                    "valeur": entry["Valeur_Mesuree"],
                    "unite": entry["Unité"],
                    "score": entry["Score_Parametre"],
                    "duree": entry["Duree"],
                    "etendue": entry["Etendue"],
                    "frequence": entry["Frequence"],
                    "amplitude": entry["Amplitude"],
                    "classification": entry["Classification"]
                }
                parametres.append(param_dict)
            
            rapports_phases[phase] = {
                "parametres": parametres,
                "score_moyen": np.mean([p["score"] for p in parametres if p["score"]]) if parametres else 0
            }
        
        # Statistiques globales
        scores_par_phase = {phase: data["score_moyen"] for phase, data in rapports_phases.items()}
        
        # Identifier les risques majeurs
        risques_majeurs = []
        for entry in scored_data:
            if entry["Classification"] in ["FORT", "TRÈS GRAVE"]:
                risques_majeurs.append({
                    "parametre": entry["Paramètre"],
                    "milieu": entry["Milieu"],
                    "amplitude": entry["Classification"],
                    "phase": entry["Phase"]
                })
        
        # Recommandations
        recommandations = self._generate_recommendations(risques_majeurs)
        
        assessment = {
            "rapports_phases": rapports_phases,
            "statistiques_globales": {
                "scores_par_phase": scores_par_phase,
                "risques_majeurs": risques_majeurs,
                "recommandations": recommandations,
                "score_global": np.mean(list(scores_par_phase.values()))
            },
            "localisation": {
                "latitude": coordinates[0],
                "longitude": coordinates[1]
            },
            "date_evaluation": datetime.now().isoformat()
        }
        
        return assessment
    
    def export_to_excel(self, assessment_data, output_path):
        """
        Exporte les résultats d'évaluation vers un fichier Excel selon le format SLRI de référence
        """
        try:
            # Charger le template SLRI
            template_path = os.path.join(self.slri_path, "Standardiser-levaluation-des-risques-et-impacts.xlsx")
            
            if os.path.exists(template_path):
                # Copier le template et le remplir
                wb = load_workbook(template_path)
                ws_matrice = wb["Matrice SLRI"]
                
                # Remplir la matrice avec les données collectées
                if "rapports_phases" in assessment_data:
                    row = 2  # Commencer après l'en-tête
                    
                    for phase_name, phase_data in assessment_data["rapports_phases"].items():
                        if "parametres" in phase_data:
                            for param in phase_data["parametres"]:
                                # Trouver la ligne correspondante dans la matrice
                                for matrix_row in range(2, ws_matrice.max_row + 1):
                                    if (ws_matrice.cell(matrix_row, 1).value == phase_name and 
                                        ws_matrice.cell(matrix_row, 2).value == param.get("milieu", "") and
                                        ws_matrice.cell(matrix_row, 3).value == param.get("parametre", "")):
                                        
                                        # Remplir les valeurs
                                        ws_matrice.cell(matrix_row, 4).value = param.get("valeur", "")  # Valeur mesurée
                                        ws_matrice.cell(matrix_row, 7).value = param.get("score", "")  # Score paramètre
                                        ws_matrice.cell(matrix_row, 8).value = param.get("duree", "")  # Durée
                                        ws_matrice.cell(matrix_row, 9).value = param.get("etendue", "")  # Étendue
                                        ws_matrice.cell(matrix_row, 10).value = param.get("frequence", "")  # Fréquence
                                        ws_matrice.cell(matrix_row, 11).value = param.get("amplitude", "")  # Amplitude
                                        ws_matrice.cell(matrix_row, 12).value = param.get("classification", "")  # Classification
                                        break
                
                # Remplir la feuille de synthèse
                if "Synthèse" in wb.sheetnames and "statistiques_globales" in assessment_data:
                    ws_synthese = wb["Synthèse"]
                    stats = assessment_data["statistiques_globales"]
                    
                    if "scores_par_phase" in stats:
                        row = 2
                        for phase, score in stats["scores_par_phase"].items():
                            ws_synthese.cell(row, 1).value = phase
                            ws_synthese.cell(row, 2).value = score
                            row += 1
                
                wb.save(output_path)
            else:
                # Créer un nouveau fichier si le template n'existe pas
                self._create_excel_from_data(assessment_data, output_path)
            
            logger.info(f"Rapport SLRI exporté vers: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export Excel: {e}")
            return False
    
    def _create_excel_from_data(self, assessment_data, output_path):
        """
        Crée un nouveau fichier Excel à partir des données si le template n'existe pas
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Feuille principale avec tous les paramètres
                if "rapports_phases" in assessment_data:
                    all_params = []
                    for phase, data in assessment_data["rapports_phases"].items():
                        if "parametres" in data:
                            for param in data["parametres"]:
                                param_row = {
                                    "Phase": phase,
                                    "Milieu": param.get("milieu", ""),
                                    "Paramètre": param.get("parametre", ""),
                                    "Valeur": param.get("valeur", ""),
                                    "Unité": param.get("unite", ""),
                                    "Score": param.get("score", 0),
                                    "Amplitude": param.get("amplitude", ""),
                                    "Classification": param.get("classification", "")
                                }
                                all_params.append(param_row)
                    
                    if all_params:
                        df_params = pd.DataFrame(all_params)
                        df_params.to_excel(writer, sheet_name="Matrice SLRI", index=False)
                
                # Feuille de synthèse
                if "statistiques_globales" in assessment_data:
                    stats = assessment_data["statistiques_globales"]
                    
                    # Scores par phase
                    if "scores_par_phase" in stats:
                        phases_data = []
                        for phase, score in stats["scores_par_phase"].items():
                            phases_data.append({"Phase": phase, "Score Moyen": score})
                        
                        if phases_data:
                            df_phases = pd.DataFrame(phases_data)
                            df_phases.to_excel(writer, sheet_name="Synthèse", index=False)
            
            return True
        except Exception as e:
            logger.error(f"Erreur lors de la création du fichier Excel: {e}")
            return False

def integrate_slri_with_main_system(coordinates, project_type="SLRI"):
    """
    Fonction d'intégration principale pour connecter SLRI au système principal
    Remplit automatiquement la matrice SLRI avec les données collectées
    """
    try:
        # Initialiser l'analyseur SLRI
        slri_dir = os.path.join(os.path.dirname(__file__), "SLRI")
        analyzer = SLRIAnalyzer(slri_dir)
        
        # Charger les données SLRI de référence
        slri_data = analyzer.load_slri_data()
        
        if not slri_data:
            return {"error": "Impossible de charger les données SLRI"}
        
        # Simuler la collecte de données environnementales pour les coordonnées
        collected_data = analyzer._simulate_data_collection(coordinates)
        
        # Mapper les données collectées sur la structure SLRI
        mapped_data = analyzer._map_data_to_slri_structure(collected_data, slri_data)
        
        # Calculer les scores selon la méthodologie SLRI
        scored_data = analyzer._calculate_slri_scores(mapped_data)
        
        # Générer l'évaluation globale selon le format SLRI
        assessment = analyzer._generate_slri_assessment(scored_data, coordinates)
        
        # Créer un fichier Excel rempli selon le format SLRI
        output_path = f"output/rapport_slri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        analyzer.export_to_excel(assessment, output_path)
        
        return assessment
        
    except Exception as e:
        logger.error(f"Erreur d'intégration SLRI: {e}")
        return {"error": str(e)}


if __name__ == "__main__":
    # Test de l'intégration SLRI
    slri_path = "SLRI"
    coordinates = (34.0209, -6.8416)  # Exemple Rabat
    
    result = integrate_slri_with_main_system(coordinates, slri_path)
    print(json.dumps(result, indent=2, ensure_ascii=False))
